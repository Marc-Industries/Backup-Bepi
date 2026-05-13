"""Excel import/export service — openpyxl-based."""
import io
from dataclasses import dataclass
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1B3A5C")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, row: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_data(ws, start_row: int, end_row: int, num_cols: int):
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _auto_width(ws, num_cols: int, max_width: int = 40):
    for col in range(1, num_cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, max_width)


# ── Product Tree Export ──────────────────────────────────────────────

def export_product_tree(nodes: list[dict], budgets: dict | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Tree"

    headers = ["Code", "Name", "Level", "Subsystem", "Parent Code", "Quantity",
               "Mass (kg)", "Mass Margin %", "Power (W)", "Power Margin %",
               "Maturity", "Manufacturer", "TRL", "Heritage", "Status"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    code_map = {n["id"]: n["code"] for n in nodes}
    for i, n in enumerate(nodes, 2):
        parent_code = code_map.get(n.get("parent_id"), "")
        b = (budgets or {}).get(n["code"], {})
        ws.cell(row=i, column=1, value=n["code"])
        ws.cell(row=i, column=2, value=n["name"])
        ws.cell(row=i, column=3, value=n.get("level", ""))
        ws.cell(row=i, column=4, value=n.get("subsystem_type", ""))
        ws.cell(row=i, column=5, value=parent_code)
        ws.cell(row=i, column=6, value=n.get("quantity", 1))
        ws.cell(row=i, column=7, value=b.get("mass"))
        ws.cell(row=i, column=8, value=b.get("mass_margin"))
        ws.cell(row=i, column=9, value=b.get("power"))
        ws.cell(row=i, column=10, value=b.get("power_margin"))
        ws.cell(row=i, column=11, value=b.get("maturity", ""))
        ws.cell(row=i, column=12, value=n.get("manufacturer", ""))
        ws.cell(row=i, column=13, value=n.get("trl"))
        ws.cell(row=i, column=14, value=n.get("heritage", ""))
        ws.cell(row=i, column=15, value=n.get("status", ""))

    _style_data(ws, 2, len(nodes) + 1, len(headers))
    _auto_width(ws, len(headers))
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Requirements Export ──────────────────────────────────────────────

def export_requirements(requirements: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"

    headers = ["Req ID", "Title", "Text", "Level", "Category",
               "Verification Method", "Verification Status",
               "Owner", "Allocated To", "Parent Req", "Priority", "Status"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    for i, r in enumerate(requirements, 2):
        ws.cell(row=i, column=1, value=r.get("req_id", ""))
        ws.cell(row=i, column=2, value=r.get("title", ""))
        ws.cell(row=i, column=3, value=r.get("text", ""))
        ws.cell(row=i, column=4, value=r.get("level", ""))
        ws.cell(row=i, column=5, value=r.get("category", ""))
        ws.cell(row=i, column=6, value=r.get("verification_method", ""))
        ws.cell(row=i, column=7, value=r.get("verification_status", ""))
        ws.cell(row=i, column=8, value=r.get("owner", ""))
        ws.cell(row=i, column=9, value=r.get("allocated_to", ""))
        ws.cell(row=i, column=10, value=r.get("parent_id", ""))
        ws.cell(row=i, column=11, value=r.get("priority", ""))
        ws.cell(row=i, column=12, value=r.get("status", ""))

    _style_data(ws, 2, len(requirements) + 1, len(headers))
    _auto_width(ws, len(headers))
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Risk Register Export ─────────────────────────────────────────────

def export_risks(risks: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Register"

    headers = ["Risk ID", "Title", "Description", "Category",
               "Likelihood", "Consequence", "Score", "Level",
               "Status", "Owner", "Mitigation"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    for i, r in enumerate(risks, 2):
        ws.cell(row=i, column=1, value=r.get("risk_id", ""))
        ws.cell(row=i, column=2, value=r.get("title", ""))
        ws.cell(row=i, column=3, value=r.get("description", ""))
        ws.cell(row=i, column=4, value=r.get("category", ""))
        ws.cell(row=i, column=5, value=r.get("likelihood"))
        ws.cell(row=i, column=6, value=r.get("consequence"))
        ws.cell(row=i, column=7, value=r.get("score"))
        ws.cell(row=i, column=8, value=r.get("level", ""))
        ws.cell(row=i, column=9, value=r.get("status", ""))
        ws.cell(row=i, column=10, value=r.get("owner", ""))
        ws.cell(row=i, column=11, value=r.get("mitigation", ""))

    _style_data(ws, 2, len(risks) + 1, len(headers))
    _auto_width(ws, len(headers))
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Schedule Export ──────────────────────────────────────────────────

def export_schedule(tasks: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    headers = ["WBS", "Task Name", "Start", "End", "Duration (days)",
               "Responsible", "Progress %", "Predecessors", "Status"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    for i, t in enumerate(tasks, 2):
        ws.cell(row=i, column=1, value=t.get("wbs", ""))
        ws.cell(row=i, column=2, value=t.get("name", ""))
        ws.cell(row=i, column=3, value=t.get("start", ""))
        ws.cell(row=i, column=4, value=t.get("end", ""))
        ws.cell(row=i, column=5, value=t.get("duration", ""))
        ws.cell(row=i, column=6, value=t.get("responsible", ""))
        ws.cell(row=i, column=7, value=t.get("progress", 0))
        ws.cell(row=i, column=8, value=", ".join(t.get("predecessors", [])))
        ws.cell(row=i, column=9, value=t.get("status", ""))

    _style_data(ws, 2, len(tasks) + 1, len(headers))
    _auto_width(ws, len(headers))
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Full Mission Export (multi-sheet) ────────────────────────────────

def export_mission(nodes: list[dict], budgets: dict, requirements: list[dict],
                   risks: list[dict], tasks: list[dict],
                   mission_info: dict | None = None) -> bytes:
    wb = Workbook()

    # Mission Info sheet
    ws_info = wb.active
    ws_info.title = "Mission"
    if mission_info:
        ws_info.cell(row=1, column=1, value="Parameter").font = Font(bold=True)
        ws_info.cell(row=1, column=2, value="Value").font = Font(bold=True)
        _style_header(ws_info, 1, 2)
        for i, (k, v) in enumerate(mission_info.items(), 2):
            ws_info.cell(row=i, column=1, value=k)
            ws_info.cell(row=i, column=2, value=str(v))
        _auto_width(ws_info, 2)

    # Product Tree sheet
    ws_tree = wb.create_sheet("Product Tree")
    tree_headers = ["Code", "Name", "Level", "Subsystem", "Parent Code", "Qty",
                    "Mass (kg)", "Power (W)", "Maturity", "TRL"]
    for c, h in enumerate(tree_headers, 1):
        ws_tree.cell(row=1, column=c, value=h)
    _style_header(ws_tree, 1, len(tree_headers))
    code_map = {n["id"]: n["code"] for n in nodes}
    for i, n in enumerate(nodes, 2):
        b = budgets.get(n["code"], {})
        ws_tree.cell(row=i, column=1, value=n["code"])
        ws_tree.cell(row=i, column=2, value=n["name"])
        ws_tree.cell(row=i, column=3, value=n.get("level", ""))
        ws_tree.cell(row=i, column=4, value=n.get("subsystem_type", ""))
        ws_tree.cell(row=i, column=5, value=code_map.get(n.get("parent_id"), ""))
        ws_tree.cell(row=i, column=6, value=n.get("quantity", 1))
        ws_tree.cell(row=i, column=7, value=b.get("mass"))
        ws_tree.cell(row=i, column=8, value=b.get("power"))
        ws_tree.cell(row=i, column=9, value=b.get("maturity", ""))
        ws_tree.cell(row=i, column=10, value=n.get("trl"))
    _style_data(ws_tree, 2, len(nodes) + 1, len(tree_headers))
    _auto_width(ws_tree, len(tree_headers))

    # Requirements sheet
    ws_req = wb.create_sheet("Requirements")
    req_headers = ["Req ID", "Title", "Text", "Level", "Category",
                   "Method", "Status", "Owner", "Allocated To"]
    for c, h in enumerate(req_headers, 1):
        ws_req.cell(row=1, column=c, value=h)
    _style_header(ws_req, 1, len(req_headers))
    for i, r in enumerate(requirements, 2):
        ws_req.cell(row=i, column=1, value=r.get("req_id", ""))
        ws_req.cell(row=i, column=2, value=r.get("title", ""))
        ws_req.cell(row=i, column=3, value=r.get("text", ""))
        ws_req.cell(row=i, column=4, value=r.get("level", ""))
        ws_req.cell(row=i, column=5, value=r.get("category", ""))
        ws_req.cell(row=i, column=6, value=r.get("verification_method", ""))
        ws_req.cell(row=i, column=7, value=r.get("verification_status", ""))
        ws_req.cell(row=i, column=8, value=r.get("owner", ""))
        ws_req.cell(row=i, column=9, value=r.get("allocated_to", ""))
    _style_data(ws_req, 2, len(requirements) + 1, len(req_headers))
    _auto_width(ws_req, len(req_headers))

    # Risks sheet
    ws_risk = wb.create_sheet("Risks")
    risk_headers = ["Risk ID", "Title", "Category", "L", "C", "Score", "Level", "Status", "Owner", "Mitigation"]
    for c, h in enumerate(risk_headers, 1):
        ws_risk.cell(row=1, column=c, value=h)
    _style_header(ws_risk, 1, len(risk_headers))
    for i, r in enumerate(risks, 2):
        ws_risk.cell(row=i, column=1, value=r.get("risk_id", ""))
        ws_risk.cell(row=i, column=2, value=r.get("title", ""))
        ws_risk.cell(row=i, column=3, value=r.get("category", ""))
        ws_risk.cell(row=i, column=4, value=r.get("likelihood"))
        ws_risk.cell(row=i, column=5, value=r.get("consequence"))
        ws_risk.cell(row=i, column=6, value=r.get("score"))
        ws_risk.cell(row=i, column=7, value=r.get("level", ""))
        ws_risk.cell(row=i, column=8, value=r.get("status", ""))
        ws_risk.cell(row=i, column=9, value=r.get("owner", ""))
        ws_risk.cell(row=i, column=10, value=r.get("mitigation", ""))
    _style_data(ws_risk, 2, len(risks) + 1, len(risk_headers))
    _auto_width(ws_risk, len(risk_headers))

    # Schedule sheet
    ws_sched = wb.create_sheet("Schedule")
    sched_headers = ["WBS", "Task", "Start", "End", "Duration", "Responsible", "Progress %"]
    for c, h in enumerate(sched_headers, 1):
        ws_sched.cell(row=1, column=c, value=h)
    _style_header(ws_sched, 1, len(sched_headers))
    for i, t in enumerate(tasks, 2):
        ws_sched.cell(row=i, column=1, value=t.get("wbs", ""))
        ws_sched.cell(row=i, column=2, value=t.get("name", ""))
        ws_sched.cell(row=i, column=3, value=t.get("start", ""))
        ws_sched.cell(row=i, column=4, value=t.get("end", ""))
        ws_sched.cell(row=i, column=5, value=t.get("duration", ""))
        ws_sched.cell(row=i, column=6, value=t.get("responsible", ""))
        ws_sched.cell(row=i, column=7, value=t.get("progress", 0))
    _style_data(ws_sched, 2, len(tasks) + 1, len(sched_headers))
    _auto_width(ws_sched, len(sched_headers))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Import ───────────────────────────────────────────────────────────

def import_requirements_from_excel(file_bytes: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower().replace(" ", "_") if h else f"col_{i}" for i, h in enumerate(rows[0])]
    results = []
    for row in rows[1:]:
        d = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row) if i < len(headers)}
        if d.get("req_id") or d.get("title"):
            results.append(d)
    return results


def import_product_tree_from_excel(file_bytes: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower().replace(" ", "_") if h else f"col_{i}" for i, h in enumerate(rows[0])]
    results = []
    for row in rows[1:]:
        d = {headers[i]: (v if v is not None else "") for i, v in enumerate(row) if i < len(headers)}
        if d.get("code") or d.get("name"):
            results.append(d)
    return results


def import_risks_from_excel(file_bytes: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower().replace(" ", "_") if h else f"col_{i}" for i, h in enumerate(rows[0])]
    results = []
    for row in rows[1:]:
        d = {headers[i]: (v if v is not None else "") for i, v in enumerate(row) if i < len(headers)}
        if d.get("risk_id") or d.get("title"):
            results.append(d)
    return results
